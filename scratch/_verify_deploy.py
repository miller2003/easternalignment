import openpyxl, re, os, json
from openpyxl import load_workbook

ROOT = r"C:\Users\samja\Desktop\site\easternalignment"
TS = os.path.join(ROOT, "src", "data", "affiliateLinks.ts")
READERS = os.path.join(ROOT, "src", "content", "readers")
XLSX = os.path.join(ROOT, "scratch", "reader-deep-links.xlsx")

# parse affiliateLinks.ts
links = {}
for m in re.finditer(r'^\s*"([^"]+)":\s*"([^"]*)"', open(TS, encoding="utf-8").read(), re.M):
    links[m.group(1)] = m.group(2)

# load xlsx mapping key -> final
wb = load_workbook(XLSX); ws = wb["深层链接回填"]
xmap = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[2]: continue
    plat, slug = r[1], r[2]
    key = slug if plat == "Kasamba" else ("purple-garden-" + slug if plat == "Purple Garden" else "keen-" + slug)
    xmap[key] = r[6]

print("=== 1) xlsx 63 keys all present in affiliateLinks.ts with a TUNE deep link? ===")
bad = []
for key, deep in xmap.items():
    v = links.get(key)
    if not v:
        bad.append(f"{key}: MISSING in affiliateLinks.ts")
    elif "bargestech.go2cloud.org/aff_c" not in v or "url=" not in v:
        bad.append(f"{key}: NOT a deep link -> {v[:60]}")
    elif key.startswith("keen-") and "offer_id=209" in v:
        bad.append(f"{key}: still offer_id=209")
print("FAIL:" , bad if bad else "NONE — all 63 OK")

print("\n=== 2) every reader article CTA resolves to a deep link (no generic/empty) ===")
problems = []
for plat in ["kasamba", "purple-garden", "keen"]:
    for f in sorted(os.listdir(os.path.join(READERS, plat))):
        if not f.endswith(".md"): continue
        txt = open(os.path.join(READERS, plat, f), encoding="utf-8").read()
        fm = re.match(r"^---\s*\n(.*?)\n---", txt, re.S)
        fmx = fm.group(1) if fm else ""
        aff = re.search(r"^affiliateUrl:\s*(.*)$", fmx, re.M)
        aff = aff.group(1).strip().strip('"') if aff else None
        unavailable = "unavailable: true" in fmx
        if unavailable:
            # unavailable readers legitimately fall back to the generic platform link
            continue
        if not aff or aff == "#":
            problems.append(f"{plat}/{f}: no affiliateUrl"); continue
        # extract slug
        mm = re.match(r"^/go/(.+?)/?$", aff)
        slug = mm.group(1) if mm else None
        if slug is None:
            problems.append(f"{plat}/{f}: weird aff {aff}"); continue
        v = links.get(slug)
        if not v or "bargestech.go2cloud.org/aff_c" not in v or "url=" not in v:
            problems.append(f"{plat}/{f}: CTA /go/{slug}/ -> NOT deep link ({ (v or 'MISSING')[:50] })")
        # inline body links must also resolve
        for href in set(re.findall(r'href="(/go/[^"]+)"', txt)):
            s2 = href.strip("/").replace("go/", "", 1) if href.startswith("/go/") else href
            s2 = s2.rstrip("/")
            vv = links.get(s2)
            if not vv or "bargestech.go2cloud.org/aff_c" not in vv or "url=" not in vv:
                problems.append(f"{plat}/{f}: inline {href} -> NOT deep link")
print("PROBLEMS:", problems if problems else "NONE — every button resolves to a deep link")

print("\n=== 3) orphan check: any /go/ key used by an article but missing in affiliateLinks.ts? ===")
used = set()
for plat in ["kasamba", "purple-garden", "keen"]:
    for f in os.listdir(os.path.join(READERS, plat)):
        if not f.endswith(".md"): continue
        txt = open(os.path.join(READERS, plat, f), encoding="utf-8").read()
        for href in set(re.findall(r'href="(/go/[^"]+)"', txt)):
            s = href.replace("/go/", "").strip("/")
            used.add(s)
missing = [u for u in used if u not in links]
print("article-referenced keys missing from affiliateLinks.ts:", missing if missing else "NONE")
print(f"\naffiliateLinks.ts total keys: {len(links)} | xlsx keys: {len(xmap)}")
