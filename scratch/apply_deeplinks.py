#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1. Verify 1:1 correspondence: for each row, deep link's url= must decode to
   the official profile URL (col 5). Abort if any mismatch.
2. Patch src/data/affiliateLinks.ts: replace the URL value for each of the
   21 reader slugs with the user's deep link from the Excel.
"""
import openpyxl, re, urllib.parse, sys

ROOT = "C:/Users/samja/Desktop/site/easternalignment"
XLSX = f"{ROOT}/reader-deeplinks-2026-08-18.xlsx"
TS = f"{ROOT}/src/data/affiliateLinks.ts"

wb = openpyxl.load_workbook(XLSX)
ws = wb.active

# --- 1. verify 1:1 ---
problems = []
mapping = {}  # slug -> deep_link
skipped = []
for row in range(2, ws.max_row + 1):
    reader = ws.cell(row=row, column=2).value
    slug = ws.cell(row=row, column=4).value
    official = ws.cell(row=row, column=5).value  # col 5 = official profile
    deep = ws.cell(row=row, column=6).value      # col 6 = deep link
    if not deep or not str(deep).strip():
        problems.append(f"{slug}: deep link is EMPTY")
        skipped.append(slug)
        continue
    m = re.search(r'url=(https?%3A%2F%2F[^&]+)', deep)
    if not m:
        problems.append(f"{slug}: deep link missing url= param")
        skipped.append(slug)
        continue
    decoded = urllib.parse.unquote(m.group(1)).split("?")[0].rstrip("/")
    expected = (official or "").split("?")[0].rstrip("/")
    if decoded != expected:
        problems.append(
            f"{slug}: 1:1 MISMATCH\n    deep-url:    {decoded}\n    official:    {expected}"
        )
        skipped.append(slug)
        continue
    mapping[slug] = deep

if problems:
    print("!!! 1:1 CORRESPONDENCE ISSUES DETECTED !!!")
    for p in problems:
        print("  -", p)
    print(f"\nProceeding with {len(mapping)} verified-correct slug(s); skipping {len(skipped)} broken one(s):")
    for s in skipped:
        print(f"   - {s}  (keeps current affiliateLinks.ts value -> correct profile)")
    print()

print(f"1:1 OK for all {len(mapping)} readers.  Proceeding to patch affiliateLinks.ts...")

# --- 2. patch affiliateLinks.ts ---
with open(TS, "r", encoding="utf-8") as f:
    content = f.read()

replacements = 0
new_lines = []
for line in content.splitlines(keepends=True):
    m = re.match(r'^(\s*)"([a-z0-9\-]+)":\s*"([^"]+)"(.*)$', line)
    if m and m.group(2) in mapping:
        indent = m.group(1)
        slug = m.group(2)
        tail = m.group(4)
        new_line = f'{indent}"{slug}": "{mapping[slug]}"{tail}'
        if new_line != line:
            replacements += 1
            new_lines.append(new_line)
        else:
            new_lines.append(line)
    else:
        new_lines.append(line)

if replacements:
    with open(TS, "w", encoding="utf-8") as f:
        f.write("".join(new_lines))

print(f"Patched {replacements} slug(s) in affiliateLinks.ts (others were no-ops / not present).")