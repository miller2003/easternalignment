from openpyxl import load_workbook
import json

path = r'C:/Users/samja/Desktop/site/easternalignment/scratch/reader-deep-links.xlsx'
wb = load_workbook(path)
ws = wb['深层链接回填']
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
print("HEADERS:", header)
out = []
for r in rows[1:]:
    if r[0] is None:
        continue
    rec = {
        'platform': r[1],
        'slug': r[2],
        'reader': r[3],
        'official': r[4],
        'offer_id': r[5],
        'deep': r[6],
    }
    out.append(rec)
    print(f"[{rec['platform']}] slug={rec['slug']} offer={rec['offer_id']}")
    print(f"    DEEP={rec['deep']}")
print("\nTOTAL ROWS:", len(out))
# dump for programmatic use
with open(r'C:/Users/samja/Desktop/site/easternalignment/scratch/_xlsx_dump.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
