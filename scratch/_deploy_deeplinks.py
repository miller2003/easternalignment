import openpyxl, urllib.parse, re, os, json, sys
from openpyxl import load_workbook

ROOT = r"C:\Users\samja\Desktop\site\easternalignment"
XLSX = os.path.join(ROOT, "scratch", "reader-deep-links.xlsx")
TS = os.path.join(ROOT, "src", "data", "affiliateLinks.ts")
READERS = os.path.join(ROOT, "src", "content", "readers")
MODE = sys.argv[1] if len(sys.argv) > 1 else "validate"
MAPJSON = os.path.join(ROOT, "scratch", "_deploy_mapping.json")

EXP_OFFER = {"Kasamba": 191, "Purple Garden": 30, "Keen": 221}
PLATDIR = {"Kasamba": "kasamba", "Purple Garden": "purple-garden", "Keen": "keen"}
PG_GENERIC = {"psychic-medium-chloe", "empathic-intuitive-marcus", "tarot-by-elena", "twin-flame-specialist-aria"}

wb = load_workbook(XLSX)
ws = wb["深层链接回填"]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[2]:
        continue
    rows.append({"plat": r[1], "slug": r[2], "official": r[4], "off_xlsx": r[5], "deep": r[6]})


def compute_key(plat, slug):
    if plat == "Kasamba":
        return slug
    if plat == "Purple Garden":
        return "purple-garden-" + slug
    if plat == "Keen":
        return "keen-" + slug


def normalize(u):
    return u.split("?")[0].split("#")[0].rstrip("/")


def parse_ts():
    txt = open(TS, encoding="utf-8").read()
    d = {}
    for m in re.finditer(r'^\s*"([^"]+)":\s*"([^"]*)"', txt, re.M):
        d[m.group(1)] = m.group(2)
    return d


mapping = {}
for row in rows:
    plat, slug, official, off_xlsx, deep = row["plat"], row["slug"], row["official"], row["off_xlsx"], row["deep"]
    key = compute_key(plat, slug)
    issues = []
    if not deep or "bargestech.go2cloud.org/aff_c" not in deep:
        issues.append("NOT_A_TUNE_LINK")
    else:
        m = re.search(r"offer_id=(\d+)", deep)
        off = int(m.group(1)) if m else None
        if plat == "Keen":
            if off == 209:
                issues.append("OFFER_209_WILL_NORMALIZE_TO_221")
            elif off != 221:
                issues.append(f"OFFER_{off}_UNEXPECTED")
        else:
            if off != EXP_OFFER[plat]:
                issues.append(f"OFFER_{off}_MISMATCH_EXPECTED_{EXP_OFFER[plat]}")
        mu = re.search(r"url=(.*?)(?:&aff_id=|&$)", deep + "&", re.S)
        if mu:
            base = normalize(urllib.parse.unquote(mu.group(1)))
            if base != normalize(official or ""):
                issues.append(f"URL_MISMATCH deep->{base} vs official->{normalize(official or '')}")
        else:
            issues.append("NO_URL_PARAM")

    if plat == "Kasamba" and slug == "kasamba-cosmic-fusion" and "bargestech.go2cloud.org/aff_c" not in (deep or ""):
        issues.append("COSMIC_FUSION_MALFORMED_USED_EXISTING")
        final = None
    elif plat == "Keen":
        final = deep.replace("offer_id=209", "offer_id=221") if "offer_id=209" in deep else deep
    else:
        final = deep

    edit_article = (plat == "Keen") or (plat == "Purple Garden" and slug in PG_GENERIC)
    mapping[key] = {"plat": plat, "slug": slug, "official": official, "deep": deep,
                    "final": final, "edit_article": edit_article, "issues": issues}

# ---- VALIDATE MODE ----
if MODE == "validate":
    print(f"Total rows: {len(rows)}  | keys: {len(mapping)}")
    print(f"{'KEY':46} {'PLAT':12} {'EDIT':4} ISSUES")
    for key, info in mapping.items():
        print(f"{key:46} {info['plat']:12} {'Y' if info['edit_article'] else 'n':4} {','.join(info['issues']) or 'OK'}")
    serious = [k for k, i in mapping.items()
               if any(x not in ("OFFER_209_WILL_NORMALIZE_TO_221", "COSMIC_FUSION_MALFORMED_USED_EXISTING") for x in i["issues"])]
    print("\nSerious issues (auto-fixed ones excluded):", serious or "NONE")
    n_edit = sum(1 for i in mapping.values() if i["edit_article"])
    print(f"Articles to edit: {n_edit} (23 Keen + 4 PG generic)")
    json.dump(mapping, open(MAPJSON, "w"), indent=2)
    sys.exit(0)

# ---- APPLY MODE ----
existing = parse_ts()
# resolve cosmic-fusion final from existing TS
for key, info in mapping.items():
    if info["final"] is None and key in existing:
        info["final"] = existing[key]

# 1) Update affiliateLinks.ts
lines = open(TS, encoding="utf-8").read().split("\n")
seen = set()
out = []
for line in lines:
    m = re.match(r'^(\s*)"([^"]+)":\s*"([^"]*)",?\s*$', line)
    if m:
        indent, k, v = m.group(1), m.group(2), m.group(3)
        if k in mapping and mapping[k]["final"] is not None:
            line = f'{indent}"{k}": "{mapping[k]["final"]}",'
            seen.add(k)
    out.append(line)

# append new keys (not already present) before closing "};"
new_keys = [k for k in mapping if k not in seen and mapping[k]["final"] is not None]
if new_keys:
    insert_at = next(i for i, l in enumerate(out) if l.strip() == "};")
    block = [f'  "{k}": "{mapping[k]["final"]}",' for k in new_keys]
    out[insert_at:insert_at] = block
open(TS, "w", encoding="utf-8").write("\n".join(out))
print(f"affiliateLinks.ts updated. New keys appended: {len(new_keys)}")

# 2) Edit article files (frontmatter + inline)
edited = 0
for key, info in mapping.items():
    if not info["edit_article"]:
        continue
    fpath = os.path.join(READERS, PLATDIR[info["plat"]], info["slug"] + ".md")
    if not os.path.exists(fpath):
        print(f"  !! MISSING ARTICLE: {fpath}")
        continue
    t = open(fpath, encoding="utf-8").read()
    if info["plat"] == "Keen":
        t = re.sub(r'affiliateUrl:\s*"/go/keen/"', f'affiliateUrl: "/go/{key}/"', t)
        t = t.replace("affiliateUrl: /go/keen/", f"affiliateUrl: /go/{key}/")
        t = t.replace('href="/go/keen/"', f'href="/go/{key}/"')
    else:  # PG generic
        t = re.sub(r'affiliateUrl:\s*"/go/purple-garden/"', f'affiliateUrl: "/go/{key}/"', t)
        t = t.replace("affiliateUrl: /go/purple-garden/", f"affiliateUrl: /go/{key}/")
    open(fpath, "w", encoding="utf-8").write(t)
    edited += 1
print(f"Articles edited: {edited}")

json.dump(mapping, open(MAPJSON, "w"), indent=2)
print("Done. Mapping saved to", MAPJSON)
