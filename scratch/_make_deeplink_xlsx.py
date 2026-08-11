# -*- coding: utf-8 -*-
"""Parse scratch/reader-official-urls.md -> scratch/reader-deep-links.xlsx
Last column '深层链接（你填）' is left empty for the user to fill.
"""
import re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"C:\Users\samja\Desktop\site\easternalignment"
SRC = os.path.join(BASE, r"scratch\reader-official-urls.md")
OUT = os.path.join(BASE, r"scratch\reader-deep-links.xlsx")

with open(SRC, encoding="utf-8") as f:
    lines = f.read().splitlines()

rows = []
platform = None
in_deleted = False
header_cols = None

def split_row(line):
    line = line.strip()
    if line.startswith("|"):
        line = line[1:]
    if line.endswith("|"):
        line = line[:-1]
    return [c.strip() for c in line.split("|")]

for s in lines:
    s = s.strip()
    if s.startswith("## 一、"):
        platform, header_cols = "Kasamba", None; in_deleted = False; continue
    if s.startswith("## 二、"):
        platform, header_cols = "Purple Garden", None; in_deleted = False; continue
    if s.startswith("## 三、"):
        platform, header_cols = "Keen", None; in_deleted = False; continue
    if s.startswith("**❌") or "已删除文章" in s:
        in_deleted = True; header_cols = None; continue
    if "已找到官方 URL" in s:
        in_deleted = False; header_cols = None; continue
    if in_deleted:
        continue
    # reset header when leaving a table block
    if not s.startswith("|"):
        header_cols = None
        continue
    # separator row
    if re.match(r'^[|:\-\s]+$', s) and '-' in s:
        continue
    cells = split_row(s)
    # header detection
    if header_cols is None and any(k in s for k in ["我的 slug", "官方链接"]):
        header_cols = {}
        for idx, c in enumerate(cells):
            if "我的 slug" in c: header_cols["slug"] = idx
            elif "解读师" in c: header_cols["advisor"] = idx
            elif "官方链接" in c: header_cols["url"] = idx
            elif "offer_id" in c: header_cols["offer"] = idx
        continue
    if header_cols is not None:
        def get(k):
            i = header_cols.get(k)
            if i is None or i >= len(cells): return ""
            return cells[i]
        slug, url, offer, advisor = get("slug"), get("url"), get("offer"), get("advisor")
        if not url:
            continue
        rows.append({
            "platform": platform or "",
            "slug": slug,
            "advisor": advisor if advisor else slug,
            "url": url,
            "offer": offer,
        })

print("parsed rows:", len(rows))

# ---- build workbook ----
wb = Workbook()

# Sheet 1: 深层链接回填
ws = wb.active
ws.title = "深层链接回填"
headers = ["#", "平台", "我的slug", "解读师", "官方链接", "offer_id", "深层链接（你填）"]
ws.append(headers)

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
fill_yellow = PatternFill("solid", fgColor="FFF2CC")  # fill-here column
center = Alignment(horizontal="center", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for i, r in enumerate(rows, start=1):
    ws.append([i, r["platform"], r["slug"], r["advisor"], r["url"], r["offer"], ""])
    rr = i + 1
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=rr, column=c)
        cell.border = border
        if c in (1, 2, 6):
            cell.alignment = center
        elif c == 7:
            cell.alignment = wrap
            cell.fill = fill_yellow
        else:
            cell.alignment = wrap

widths = [5, 13, 26, 22, 62, 9, 50]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{len(rows) + 1}"

# Sheet 2: 说明
ws2 = wb.create_sheet("说明")
notes = [
    ("用法说明", True),
    ("", False),
    ("1. 本表「深层链接（你填）」列（末列，黄色）为空，由你填写。", False),
    ("2. 拿每行的「官方链接」去 TUNE / Barges 后台生成深层链接（公式见下），把生成的完整链接粘贴到该行末列。", False),
    ("3. 填完把本文件发回给我，我批量替换 src/data/affiliateLinks.ts 对应条目并 rebuild。", False),
    ("", False),
    ("深层链接公式：", True),
    ("https://bargestech.go2cloud.org/aff_c?offer_id=<offer_id>&aff_id=2326&url=<官方链接 经 URL 编码>", False),
    ("", False),
    ("offer_id 映射：  Kasamba = 191   |   Purple Garden = 30   |   Keen = 221", False),
    ("", False),
    ("重要 — 验证状态（2026-08-11）：", True),
    ("• Kasamba 20 + Purple Garden 20 = 40 条：已实测 HTTP 200，确定有效，可直接建。", False),
    ("• Keen 23 条：今天我方环境被 Cloudflare 封锁（curl 与 WebFetch 均打不开），无法复验；", False),
    ("  这些链接是 2026-08-10 审计时验证过的。建深层链接前请在本机逐条点开确认能打开。", False),
    ("• 优先检查 Keen 这 4 条结构最脆弱的：master-psychic-dev（http://）、intuitive-jade（http:// 且真实 slug 是 isis-jade）、", False),
    ("  ladyfontaine、lollie-ext-5555（无数字 ID 的非标准 Keen URL）。", False),
    ("", False),
    ("哪条打不开，把对应「我的slug」发我，我来处理（换 URL 或删文）。", False),
]
for i, (text, bold) in enumerate(notes, start=1):
    cell = ws2.cell(row=i, column=1, value=text)
    cell.font = Font(bold=bold, size=12 if bold else 11, color="1F4E78" if bold else "000000")
    cell.alignment = Alignment(wrap_text=True, vertical="center")
ws2.column_dimensions["A"].width = 110

wb.save(OUT)
print("saved:", OUT)
