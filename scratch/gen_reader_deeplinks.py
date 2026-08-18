#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate an Excel mapping of completed reader reviews -> their /go/ slugs,
current affiliate (TUNE) base URL, and a blank column for the user to paste
deep links generated in the affiliate backend.

Also reports which affiliate slugs have NO corresponding article file
(the "unfinished" part of the interrupted batch).
"""
import os, re, glob, json, urllib.parse

ROOT = "C:/Users/samja/Desktop/site/easternalignment"
READERS_DIR = os.path.join(ROOT, "src", "content", "readers")
AFF_TS = os.path.join(ROOT, "src", "data", "affiliateLinks.ts")
OUT_XLSX = os.path.join(ROOT, "reader-deeplinks-2026-08-18.xlsx")

# --- 1. Parse affiliateLinks.ts : slug -> url ---
slug_url = {}
with open(AFF_TS, "r", encoding="utf-8") as f:
    for line in f:
        m = re.search(r'"([a-z0-9\-]+)":\s*"([^"]+)"', line)
        if m:
            slug_url[m.group(1)] = m.group(2)

# --- 2. Parse all reader md files: collect (slug -> file) and frontmatter ---
def parse_frontmatter(text):
    if not text.startswith("---"):
        return {}
    # frontmatter between first and second '---'
    end = text.find("\n---", 3)
    if end == -1:
        return {}
    fm = text[3:end]
    data = {}
    for line in fm.splitlines():
        if re.match(r"^\s+\S", line):  # indented -> nested, skip
            continue
        m = re.match(r'^([A-Za-z_]+):\s?(.*)$', line)
        if m:
            k, v = m.group(1), m.group(2).strip()
            v = v.strip('"').strip("'")
            data[k] = v
    return data

used_slug_to_file = {}   # slug -> md filename
file_rows = []           # list of dicts for completed articles

def decode_official(aff):
    """Extract the clean official platform profile URL from a TUNE aff_c link."""
    if not aff:
        return ""
    m = re.search(r'url=(https%3A%2F%2F[^&]+)', aff)
    if not m:
        return ""
    decoded = urllib.parse.unquote(m.group(1))
    return decoded.split("?")[0]  # drop tracking query string

for path in glob.glob(os.path.join(READERS_DIR, "**", "*.md"), recursive=True):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    fm = parse_frontmatter(text)
    aff = fm.get("affiliateUrl", "")
    slug = aff.replace("/go/", "").strip("/")
    if slug:
        used_slug_to_file.setdefault(slug, os.path.basename(path))
        # capture only the 21 batch files (untracked / this session) by publishDate 2026-08-18
        pd = fm.get("publishDate", "")
        if "2026-08-18" in pd:
            platform = fm.get("platform", "")
            pname = fm.get("platformName", "")
            # derive clean reader display name
            if ":" in pname:
                reader = pname.split(":", 1)[1].strip()
            elif platform and slug.startswith(platform + "-"):
                reader = slug[len(platform) + 1:].replace("-", " ").title()
            else:
                reader = slug
            reader = reader.title()  # normalize casing (e.g. SYMONNE -> Symonne)
            canon = fm.get("canonicalUrl", "")
            aff = slug_url.get(slug, "")
            file_rows.append({
                "reader": reader,
                "platform": platform,
                "slug": slug,
                "official": decode_official(aff),
                "article": canon,
                "aff_url": aff,
                "file": os.path.basename(path),
            })

# order: by platform group then reader
order = {"keen": 0, "kasamba": 1, "purple-garden": 2}
file_rows.sort(key=lambda r: (order.get(r["platform"], 9), r["reader"].lower()))

# --- 3. Orphan slugs: added THIS session (git diff) but no article references them ---
import subprocess
diff = subprocess.run(
    ["git", "diff", "src/data/affiliateLinks.ts"],
    cwd=ROOT, capture_output=True, text=True
).stdout
added_this_session = set(re.findall(r'^\+\s*"([a-z0-9\-]+)":', diff, re.M))
# only count as "unfinished" if no md file anywhere references the slug
orphans = sorted(s for s in added_this_session if s not in used_slug_to_file)

print("=== COMPLETED ARTICLES (this session, publishDate 2026-08-18) ===")
print(f"count = {len(file_rows)}")
for r in file_rows:
    print(f"  [{r['platform']:13}] {r['reader']:28} slug={r['slug']}")

print("\n=== ORPHAN SLUGS (in affiliateLinks.ts, no article file) ===")
print(f"count = {len(orphans)}")
for s in sorted(orphans):
    print(f"  {s}")

# --- 4. Write xlsx ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Readers"

headers = ["#", "解读师 (Reader)", "平台 (Platform)", "/go/ Slug",
           "官网对应链接 (Official Profile URL)", "站内文章链接 (Article URL)",
           "当前联盟基础链接 (Current Affiliate URL)", "深层链接 (Deep Link) — 待填"]
ws.append(headers)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    cell.border = border

for i, r in enumerate(file_rows, start=1):
    ws.append([
        i,
        r["reader"],
        {"keen": "Keen", "kasamba": "Kasamba", "purple-garden": "Purple Garden"}.get(r["platform"], r["platform"]),
        r["slug"],
        r["official"],
        r["article"],
        r["aff_url"],
        "",  # deep link - blank for user
    ])
    row = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=row, column=c).border = border
        ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)

# column widths
widths = [4, 26, 16, 34, 56, 52, 60, 40]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"

wb.save(OUT_XLSX)
print(f"\nWrote Excel -> {OUT_XLSX}  (rows={len(file_rows)})")
